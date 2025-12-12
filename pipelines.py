import scrapy
import re
import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from scrapy.exceptions import DropItem
from .items import BookItem, CategoryItem 


# Pipeline 1: Nettoyage et conversion du prix
class PriceCleaningPipeline:
    def process_item(self, item, spider):
        if isinstance(item, BookItem):
            if item.get('price'):
 
                price_str = item['price'].replace('£', '').strip()
                try:
                    item['price_float'] = float(price_str)
                except ValueError:
                    spider.logger.warning(f"Impossible de convertir le prix '{item['price']}' en float.")
                    item['price_float'] = 0.0
            else:
                item['price_float'] = 0.0
        return item

# Pipeline 2: Dédoublonnage et comptage par catégorie
class DuplicatesAndCategoryCounterPipeline:
    def __init__(self):
        self.upc_seen = set()  
        self.category_book_counts = {} 
        self.category_names_seen = set() 

    def process_item(self, item, spider):
        if isinstance(item, BookItem):
            upc = item.get('upc')
            if upc in self.upc_seen:
                raise DropItem(f"Livre en double trouvé: {upc}")
            self.upc_seen.add(upc)

            category = item.get('category')
            if category:
                self.category_book_counts[category] = self.category_book_counts.get(category, 0) + 1
            return item
        
        if isinstance(item, CategoryItem):
            category_name = item.get('category_name')
            if category_name in self.category_names_seen:
                raise DropItem(f"Catégorie en double: {category_name}")
            self.category_names_seen.add(category_name)
            return item
        
        return item

# Pipeline 3: Export Multi-Format (JSON et Excel)
class MultiFormatExportPipeline:

    EXCEL_HEADERS = ['category', 'title', 'price', 'price_float', 'review_rating',
                     'upc', 'number_of_reviews', 'availability', 'description', 'url', 'image_url']

    def __init__(self):
        self.items_list = []
        self.category_items_list = []

    def process_item(self, item, spider):
        if isinstance(item, BookItem):
            self.items_list.append(item)
        elif isinstance(item, CategoryItem):
            self.category_items_list.append(item)
        return item

    def close_spider(self, spider):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        output_dir = 'exports'
    

        book_counts = {}

        try:
  
            for pipe in spider.crawler.engine.scraper.slot.open_pipelines:
                if isinstance(pipe, DuplicatesAndCategoryCounterPipeline):
                    book_counts = pipe.category_book_counts
                    break
        except AttributeError:
 
             spider.logger.warning("Impossible d'accéder aux autres pipelines. Le comptage par catégorie sera vide.")
             

        json_data = [dict(item) for item in self.items_list] + [dict(item) for item in self.category_items_list]
        json_filename = os.path.join(output_dir, f'export.json')
        
        try:
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=4)
            spider.logger.info(f"Données archivées au format JSON: {json_filename}")
        except Exception as e:
            spider.logger.error(f"Erreur lors de l'export JSON : {e}")


# 4. Export Excel (pour le rapport) ---
        excel_filename = os.path.join(output_dir, f'export.xlsx')
        wb = openpyxl.Workbook()

        ws_books = wb.active
        ws_books.title = "Livres"

        ws_books.append(self.EXCEL_HEADERS)

        for item in self.items_list:
            row = [item.get(header, '') for header in self.EXCEL_HEADERS]
            ws_books.append(row)

        for col in range(1, len(self.EXCEL_HEADERS) + 1):
            ws_books.column_dimensions[get_column_letter(col)].width = 20

        ws_categories = wb.create_sheet(title="Catégories")
        ws_categories.append(["Nom de Catégorie", "Nombre de Livres"])

        for category, count in book_counts.items():
            ws_categories.append([category, count])

        try:
            wb.save(excel_filename)
            spider.logger.info(f"✅ Données exportées vers Excel: {excel_filename}")
        except Exception as e:
            spider.logger.error(f"Erreur lors de l'export Excel : {e}")